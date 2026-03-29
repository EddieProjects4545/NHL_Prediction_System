"""
NHL Betting Model — Historical Backtesting & Season Analysis

Evaluates model predictive performance across:
  - 2024-25 season  (out-of-sample: model trained on 2025-26 data)
  - 2025-26 season  (YTD: in-sample, held-out val set gives honest estimate)

Since historical odds are not stored, moneyline betting P&L is simulated
at standard -110 juice for the predicted side. A bet is placed when the
model's implied edge exceeds --min-edge percent.

Usage:
  python backtest.py                    # Full report
  python backtest.py --min-edge 5       # Stricter betting filter
  python backtest.py --export excel     # Save backtest_results.xlsx
"""
import argparse
import csv
import os
import sys
from datetime import date
from typing import Dict, List, Optional, Tuple

# Force UTF-8 on Windows to handle box-drawing characters
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

import joblib
import numpy as np
import pandas as pd
from colorama import Fore, Style, init
from sklearn.metrics import (
    accuracy_score, brier_score_loss, log_loss, roc_auc_score,
)

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
init(autoreset=True)

from config import (
    CURRENT_SEASON, PREV_SEASON, GAME_TYPE_REGULAR,
    MIN_EDGE_PCT, MIN_GAMES_FOR_FEAT, SAVED_MODELS_DIR, OUTPUT_DIR,
)

# Standard -110 juice simulation constants
_SIM_ODDS    = -110
_IMPL_PROB   = 110 / (110 + 100)   # ≈ 52.38%  (implied prob at -110)
_WIN_UNITS   = round(100 / 110, 4)  # +0.9091u per winning bet


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def parse_args():
    p = argparse.ArgumentParser(description="NHL Model Backtester")
    p.add_argument("--min-edge", type=float, default=MIN_EDGE_PCT,
                   help="Min edge %% to count as a simulated bet (default %(default)s)")
    p.add_argument("--export", type=str, default="none",
                   choices=["excel", "none"], help="Export format")
    return p.parse_args()


# ---------------------------------------------------------------------------
# Metadata extraction (mirrors build_training_matrix filter logic)
# ---------------------------------------------------------------------------

def _extract_metadata(game_results: List[Dict],
                       context_results: List[Dict],
                       season_label: str) -> List[Dict]:
    """
    Replay build_training_matrix's game-skipping logic to produce a metadata
    list aligned 1-to-1 with the returned X rows.
    """
    from features.form_features import build_team_game_logs

    logs = build_team_game_logs(game_results + context_results)
    meta = []

    for g in sorted(game_results, key=lambda x: x.get("date", "")):
        home = g.get("home_team")
        away = g.get("away_team")
        gid  = g.get("game_id")
        d    = g.get("date", "")

        if not home or not away:
            continue

        h_pre = [x for x in logs.get(home, [])
                 if x["date"] < d or (x["date"] == d and x["game_id"] != gid)]
        a_pre = [x for x in logs.get(away, [])
                 if x["date"] < d or (x["date"] == d and x["game_id"] != gid)]

        if len(h_pre) < MIN_GAMES_FOR_FEAT or len(a_pre) < MIN_GAMES_FOR_FEAT:
            continue

        meta.append({
            "date":       d,
            "month":      d[:7],
            "home":       home,
            "away":       away,
            "home_goals": int(g.get("home_goals", 0)),
            "away_goals": int(g.get("away_goals", 0)),
            "season":     season_label,
        })

    return meta


# ---------------------------------------------------------------------------
# Simulation helpers
# ---------------------------------------------------------------------------

def _sim_ml_bet(model_prob: float, actual_home_win: int,
                min_edge: float) -> Tuple[Optional[str], float]:
    """
    Simulate a moneyline bet at -110. Returns (side, units) or (None, 0).
    """
    home_edge = model_prob - _IMPL_PROB
    away_edge = (1.0 - model_prob) - _IMPL_PROB

    if home_edge >= min_edge / 100:
        won   = actual_home_win == 1
        units = _WIN_UNITS if won else -1.0
        return "home", units
    if away_edge >= min_edge / 100:
        won   = actual_home_win == 0
        units = _WIN_UNITS if won else -1.0
        return "away", units

    return None, 0.0


def _sim_pl_bet(pl_prob_home: float, pl_prob_away: float,
                actual_pl_home: int, actual_pl_away: int,
                min_edge: float) -> Tuple[Optional[str], float]:
    """Simulate puck-line bet at -110."""
    home_edge = pl_prob_home - _IMPL_PROB
    away_edge = pl_prob_away - _IMPL_PROB

    if home_edge >= min_edge / 100:
        won   = actual_pl_home == 1
        return "home_pl", _WIN_UNITS if won else -1.0
    if away_edge >= min_edge / 100:
        won   = actual_pl_away == 1
        return "away_pl", _WIN_UNITS if won else -1.0

    return None, 0.0


def _calibration_table(probs: np.ndarray, actuals: np.ndarray,
                        n_bins: int = 10) -> List[Dict]:
    edges = np.linspace(0.0, 1.0, n_bins + 1)
    rows  = []
    for i in range(n_bins):
        lo, hi = edges[i], edges[i + 1]
        mask = (probs >= lo) & (probs < hi)
        n    = int(mask.sum())
        if n < 8:
            continue
        rows.append({
            "bucket":      f"{lo:.0%}–{hi:.0%}",
            "n":           n,
            "pred_avg":    round(float(probs[mask].mean()), 3),
            "actual_rate": round(float(actuals[mask].mean()), 3),
            "diff":        round(float(actuals[mask].mean() - probs[mask].mean()), 3),
        })
    return rows


# ---------------------------------------------------------------------------
# Printing helpers
# ---------------------------------------------------------------------------

_W = 57  # separator width

def _sep():
    print(Fore.WHITE + "  " + "─" * _W)

def _header(title: str, color=Fore.CYAN):
    print()
    print(color + Style.BRIGHT + f"  {'─'*3}  {title}")
    print(Fore.WHITE + "  " + "─" * _W)

def _row(label: str, value: str, color=Fore.WHITE, bold=False):
    font = Style.BRIGHT if bold else ""
    print(Fore.WHITE + f"  {label:<32}" + color + font + f"{value}")

def _fmt_pct(v: float) -> str:
    return f"{v*100:.2f}%"

def _fmt_u(v: float, compact: bool = False) -> str:
    s = f"{v:+.2f}u"
    if not compact and v != 0:
        s += f"  (${v * 10:+.2f})"
    return s


# ---------------------------------------------------------------------------
# Main report
# ---------------------------------------------------------------------------

def _season_block(label: str,
                  probs: np.ndarray, actuals: np.ndarray,
                  pl_h_probs: np.ndarray, y_pl_h: np.ndarray,
                  pl_a_probs: np.ndarray, y_pl_a: np.ndarray,
                  mu_h: np.ndarray, mu_a: np.ndarray,
                  y_gh: np.ndarray, y_ga: np.ndarray,
                  meta: List[Dict],
                  train_val_metrics: Optional[Dict],
                  min_edge: float) -> Dict:
    """
    Print one season's block and return a results dict for Excel export.
    """
    months   = [m["month"] for m in meta]
    n        = len(actuals)

    # ── Core model metrics ────────────────────────────────────────────────────
    brier    = brier_score_loss(actuals, probs)
    auc      = roc_auc_score(actuals, probs) if len(set(actuals)) > 1 else 0.5
    acc      = accuracy_score(actuals, (probs >= 0.5).astype(int))
    ll       = log_loss(actuals, probs)

    _header(label, Fore.CYAN)
    _row("Games evaluated", str(n))

    if train_val_metrics:
        # Show val-set metrics from trainer (honest 15% hold-out)
        _row("─── Validation set (last 15%)", "", bold=True)
        _row("  Brier score (lower=better)", f"{train_val_metrics['ensemble_val_brier']:.4f}")
        _row("  AUC-ROC    (0.5=random)",    f"{train_val_metrics['ensemble_val_auc']:.4f}")
        _row("  Accuracy   (0.5=random)",
             f"{train_val_metrics['ensemble_val_accuracy']:.4f}  "
             f"({(train_val_metrics['ensemble_val_accuracy']-0.5)*100:+.1f}pp vs coin flip)")
        _row("  Val set size",
             f"{train_val_metrics['n_val']} games  |  "
             f"Trained on {train_val_metrics['n_train']} games")
        print()
        _row("─── Full season (in-sample)", "", bold=True)
    else:
        _row("─── Out-of-sample metrics", "", bold=True)
        print(Fore.WHITE + "  " + Fore.YELLOW +
              "  (model trained on 2025-26; 2024-25 is truly out-of-sample)")

    _row("  Brier score", f"{brier:.4f}  (random = 0.2500)")
    _row("  AUC-ROC",     f"{auc:.4f}  (random = 0.5000)")
    _row("  Accuracy",    f"{acc:.4f}  ({(acc-0.5)*100:+.1f}pp vs coin flip)")
    _row("  Log-loss",    f"{ll:.4f}")

    # ── Home team win rate ────────────────────────────────────────────────────
    home_win_rate = float(actuals.mean())
    _row("  Home win rate (actual)", f"{home_win_rate:.1%}")
    _row("  Home win rate (model)",  f"{float(probs.mean()):.1%}")

    # ── Puck line accuracy ────────────────────────────────────────────────────
    pl_h_acc = accuracy_score(y_pl_h, (pl_h_probs >= 0.5).astype(int))
    _row("  Home -1.5 accuracy",     f"{pl_h_acc:.4f}")
    _row("  Puck line home Brier",   f"{brier_score_loss(y_pl_h, pl_h_probs):.4f}")

    # ── Poisson total accuracy ────────────────────────────────────────────────
    pred_totals   = mu_h + mu_a
    actual_totals = y_gh + y_ga
    total_mae     = float(np.abs(pred_totals - actual_totals).mean())
    _row("  Goals total MAE",        f"{total_mae:.3f} goals")
    _row("  Avg pred total",         f"{pred_totals.mean():.2f}  vs  actual {actual_totals.mean():.2f}")

    # ── Simulated betting P&L (ML @ -110) ────────────────────────────────────
    sim_bets   = []
    sim_pl_bets= []
    for i, (prob, act, m) in enumerate(zip(probs, actuals, meta)):
        side, units = _sim_ml_bet(prob, act, min_edge)
        if side:
            sim_bets.append({
                "date":   m["date"],
                "month":  m["month"],
                "game":   f"{m['away']} @ {m['home']}",
                "side":   side,
                "prob":   round(prob, 4),
                "actual": int(act),
                "units":  units,
            })
        # PL bets
        ps, pu = _sim_pl_bet(
            pl_h_probs[i], pl_a_probs[i],
            y_pl_h[i], y_pl_a[i], min_edge,
        )
        if ps:
            sim_pl_bets.append({
                "date": m["date"], "month": m["month"],
                "game": f"{m['away']} @ {m['home']}",
                "side": ps, "units": pu,
            })

    print()
    _row("─── Simulated Betting (ML @ -110, no vig arb)", "", bold=True)
    _row("  Edge threshold",  f"{min_edge:.1f}%")
    if sim_bets:
        n_bets  = len(sim_bets)
        n_wins  = sum(1 for b in sim_bets if b["units"] > 0)
        n_loss  = sum(1 for b in sim_bets if b["units"] < 0)
        total_u = sum(b["units"] for b in sim_bets)
        win_pct = n_wins / n_bets if n_bets > 0 else 0
        roi     = total_u / n_bets if n_bets > 0 else 0
        u_color = Fore.GREEN if total_u >= 0 else Fore.RED
        _row("  Bets placed",    f"{n_bets}")
        _row("  W / L",          f"{n_wins} / {n_loss}  ({win_pct:.1%} win rate)")
        _row("  Total units",    _fmt_u(total_u), color=u_color, bold=True)
        _row("  ROI per bet",    f"{roi*100:+.2f}%", color=u_color)
        _row("  Bet rate",       f"{n_bets / n:.1%} of games flagged")
    else:
        _row("  No qualifying bets", f"(raise --min-edge or lower threshold)")

    if sim_pl_bets:
        n_pl = len(sim_pl_bets)
        pl_u = sum(b["units"] for b in sim_pl_bets)
        n_pw = sum(1 for b in sim_pl_bets if b["units"] > 0)
        pl_color = Fore.GREEN if pl_u >= 0 else Fore.RED
        print()
        _row("  Puck Line bets placed", f"{n_pl}")
        _row("  PL W / L",    f"{n_pw} / {n_pl - n_pw}  ({n_pw/n_pl:.1%} win rate)")
        _row("  PL total units", _fmt_u(pl_u), color=pl_color, bold=True)

    # ── Monthly breakdown ─────────────────────────────────────────────────────
    month_arr = np.array(months)
    print()
    _row("─── Month-by-Month (ML)", "", bold=True)
    print(Fore.WHITE + f"  {'Month':<9} {'Games':>6} {'Acc':>7} {'AUC':>7} "
          f"{'Bets':>5} {'Units':>10}")
    _sep()

    for m in sorted(set(months)):
        mask = month_arr == m
        if mask.sum() < 5:
            continue
        p  = probs[mask]
        a  = actuals[mask]
        ng = int(mask.sum())
        ac = accuracy_score(a, (p >= 0.5).astype(int))
        au = roc_auc_score(a, p) if len(set(a.tolist())) > 1 else 0.5
        # Monthly bets
        m_bets  = [b for b in sim_bets if b["month"] == m]
        mb_u    = sum(b["units"] for b in m_bets)
        nb      = len(m_bets)
        u_col   = Fore.GREEN if mb_u > 0 else (Fore.RED if mb_u < 0 else Fore.WHITE)
        print(
            Fore.WHITE + f"  {m:<9} {ng:>6} {ac:>7.3f} {au:>7.4f} "
            f"{nb:>5}" + u_col + f"  {mb_u:>+7.2f}u"
        )

    # ── Calibration ───────────────────────────────────────────────────────────
    cal = _calibration_table(probs, actuals)
    print()
    _row("─── Calibration (predicted prob vs actual win rate)", "", bold=True)
    print(Fore.WHITE + f"  {'Bucket':<12} {'N':>5} {'Pred':>7} {'Actual':>7} {'Diff':>8}")
    _sep()
    for row in cal:
        diff     = row["diff"]
        d_color  = Fore.GREEN if diff > 0.02 else (Fore.RED if diff < -0.02 else Fore.WHITE)
        diff_str = f"{diff:+.3f}"
        print(
            Fore.WHITE + f"  {row['bucket']:<12} {row['n']:>5} "
            f"{row['pred_avg']:>7.3f} {row['actual_rate']:>7.3f}" +
            d_color + f" {diff_str:>8}"
        )

    return {
        "label":          label,
        "n_games":        n,
        "brier":          round(brier, 4),
        "auc":            round(auc, 4),
        "accuracy":       round(acc, 4),
        "home_win_rate":  round(home_win_rate, 4),
        "ml_bets":        sim_bets,
        "pl_bets":        sim_pl_bets,
        "total_units_ml": sum(b["units"] for b in sim_bets),
        "total_units_pl": sum(b["units"] for b in sim_pl_bets),
        "calibration":    cal,
        "train_metrics":  train_val_metrics or {},
    }


# ---------------------------------------------------------------------------
# Risk metrics
# ---------------------------------------------------------------------------

def _compute_risk_metrics(bets: List[Dict]) -> Dict:
    """Compute max drawdown and per-bet Sharpe ratio from a sorted bet list."""
    if not bets:
        return {"max_drawdown": 0.0, "sharpe": 0.0, "n_bets": 0}
    units = [b["units"] for b in sorted(bets, key=lambda x: x["date"])]
    cum = np.cumsum(units)
    running_max = np.maximum.accumulate(cum)
    drawdowns = running_max - cum
    max_dd = float(drawdowns.max())
    mean_u = float(np.mean(units))
    std_u = float(np.std(units, ddof=1)) if len(units) > 1 else 1.0
    sharpe = mean_u / std_u if std_u > 0 else 0.0
    return {"max_drawdown": round(max_dd, 4), "sharpe": round(sharpe, 4), "n_bets": len(bets)}


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

def _export_excel(results_curr: Dict, results_prev: Dict,
                  live_log: List[Dict], output_dir: str) -> str:
    """Build backtest_results.xlsx with per-season and combined analysis."""
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    NAVY  = PatternFill(fill_type="solid", fgColor="1F4E79")
    BLUE  = PatternFill(fill_type="solid", fgColor="2F75B6")
    GREEN = PatternFill(fill_type="solid", fgColor="C6EFCE")
    RED   = PatternFill(fill_type="solid", fgColor="FFCCCC")
    AMBER = PatternFill(fill_type="solid", fgColor="FFEB9C")

    W_FONT = Font(bold=True, color="FFFFFF")
    HDR    = Font(bold=True, color="FFFFFF")
    G_FONT = Font(bold=True, color="006100")
    R_FONT = Font(bold=True, color="9C0006")
    B_FONT = Font(bold=True, color="7D6608")

    wb = openpyxl.Workbook()

    def _hdr_row(ws, row, cols, fill=BLUE):
        for ci, val in enumerate(cols, 1):
            c = ws.cell(row=row, column=ci, value=val)
            c.fill = fill; c.font = HDR
            c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = 16

    def _title(ws, text, ncols=8):
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
        c = ws.cell(row=1, column=1, value=text)
        c.fill = NAVY; c.font = Font(bold=True, color="FFFFFF", size=13)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 26

    # ── Sheet 1: Summary ──────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    _title(ws, "NHL MODEL — BACKTESTING SUMMARY", ncols=8)

    sum_hdrs = ["Season", "Games", "Brier", "AUC", "Accuracy",
                "ML Bets", "ML Units", "$ P/L ($10/u)"]
    _hdr_row(ws, 2, sum_hdrs)

    for ri, res in enumerate([results_prev, results_curr], 3):
        n_bets = len(res["ml_bets"])
        units  = res["total_units_ml"]
        vals   = [
            res["label"], res["n_games"],
            res["brier"], res["auc"], f"{res['accuracy']:.1%}",
            n_bets,
            f"{units:+.2f}u",
            f"${units * 10:+.2f}",
        ]
        for ci, val in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.alignment = Alignment(horizontal="center", vertical="center")
        # Color units cell
        u_cell = ws.cell(row=ri, column=7)
        u_cell.font = G_FONT if units >= 0 else R_FONT

    # Combined row
    all_bets  = results_curr["ml_bets"] + results_prev["ml_bets"]
    all_units = sum(b["units"] for b in all_bets)
    cmb = ["COMBINED",
           results_curr["n_games"] + results_prev["n_games"],
           "", "", "",
           len(all_bets),
           f"{all_units:+.2f}u",
           f"${all_units * 10:+.2f}"]
    for ci, val in enumerate(cmb, 1):
        c = ws.cell(row=5, column=ci, value=val)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=5, column=7).font = G_FONT if all_units >= 0 else R_FONT

    # Validation metrics section
    ws.cell(row=7, column=1, value="2025-26 Validation Set (Held-out 15%)").font = Font(bold=True)
    tm = results_curr.get("train_metrics", {})
    vm_rows = [
        ("Brier score",  tm.get("ensemble_val_brier", "")),
        ("AUC-ROC",      tm.get("ensemble_val_auc",   "")),
        ("Accuracy",     tm.get("ensemble_val_accuracy", "")),
        ("Val set size", f"{tm.get('n_val','')} games"),
    ]
    for ri, (k, v) in enumerate(vm_rows, 8):
        ws.cell(row=ri, column=1, value=k)
        ws.cell(row=ri, column=2, value=v).alignment = Alignment(horizontal="center")

    # Risk metrics section
    all_bets_for_risk = results_curr["ml_bets"] + results_prev["ml_bets"]
    risk = _compute_risk_metrics(all_bets_for_risk)
    ws.cell(row=14, column=1, value="Risk Metrics (Combined ML)").font = Font(bold=True)
    risk_rows = [
        ("Max Drawdown",  f"{risk['max_drawdown']:+.2f}u"),
        ("Per-bet Sharpe", f"{risk['sharpe']:.4f}"),
    ]
    for ri, (k, v) in enumerate(risk_rows, 15):
        ws.cell(row=ri, column=1, value=k)
        ws.cell(row=ri, column=2, value=v).alignment = Alignment(horizontal="center")

    for ci, w in enumerate([14, 10, 10, 10, 12, 10, 12, 14], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # ── Sheet 2: Bet Log (combined) ───────────────────────────────────────────
    bl = wb.create_sheet("Bet Log")
    _title(bl, "NHL MODEL — SIMULATED BET LOG (ML @ -110)", ncols=7)
    _hdr_row(bl, 2, ["DATE", "SEASON", "GAME", "SIDE", "PRED%", "RESULT", "UNITS"])

    all_sorted = sorted(all_bets, key=lambda x: x["date"])
    cumulative = 0.0
    for ri, b in enumerate(all_sorted, 3):
        cumulative += b["units"]
        won = b["units"] > 0
        result_str = "WIN" if won else "LOSS"
        vals = [
            b["date"],
            b.get("season", ""),
            b["game"],
            b["side"],
            f"{b['prob']:.1%}",
            result_str,
            f"{b['units']:+.3f}u",
        ]
        for ci, val in enumerate(vals, 1):
            c = bl.cell(row=ri, column=ci, value=val)
            c.fill = GREEN if won else RED
            c.alignment = Alignment(horizontal="center" if ci != 3 else "left",
                                    vertical="center")
        bl.row_dimensions[ri].height = 14

    if all_sorted:
        tr = len(all_sorted) + 3
        bl.merge_cells(start_row=tr, start_column=1, end_row=tr, end_column=6)
        tc = bl.cell(row=tr, column=1, value="SEASON TOTAL")
        tc.font = Font(bold=True, size=11)
        tc.alignment = Alignment(horizontal="right", vertical="center")
        uc = bl.cell(row=tr, column=7,
                     value=f"{sum(b['units'] for b in all_sorted):+.2f}u")
        uc.font = G_FONT if cumulative >= 0 else R_FONT
        uc.font = Font(bold=True,
                       color="006100" if cumulative >= 0 else "9C0006",
                       size=11)
        uc.alignment = Alignment(horizontal="center", vertical="center")

    for ci, w in enumerate([12, 9, 32, 9, 8, 9, 10], 1):
        bl.column_dimensions[get_column_letter(ci)].width = w

    # ── Sheet 3: Calibration ──────────────────────────────────────────────────
    cs = wb.create_sheet("Calibration")
    _title(cs, "NHL MODEL — PROBABILITY CALIBRATION", ncols=6)

    for col_off, res in enumerate([results_prev, results_curr]):
        c_start = col_off * 7 + 1
        cs.merge_cells(start_row=2, start_column=c_start,
                        end_row=2, end_column=c_start + 4)
        hc = cs.cell(row=2, column=c_start, value=res["label"])
        hc.fill = NAVY; hc.font = W_FONT
        hc.alignment = Alignment(horizontal="center", vertical="center")

        for ci, h in enumerate(["Bucket", "N", "Pred", "Actual", "Diff"], c_start):
            c = cs.cell(row=3, column=ci, value=h)
            c.fill = BLUE; c.font = HDR
            c.alignment = Alignment(horizontal="center", vertical="center")

        for ri, row in enumerate(res["calibration"], 4):
            diff = row["diff"]
            fill = GREEN if diff > 0.02 else (RED if diff < -0.02 else AMBER)
            vals = [row["bucket"], row["n"], row["pred_avg"],
                    row["actual_rate"], f"{diff:+.3f}"]
            for ci, val in enumerate(vals, c_start):
                c = cs.cell(row=ri, column=ci, value=val)
                c.fill = fill if ci == c_start + 4 else PatternFill()
                c.alignment = Alignment(horizontal="center", vertical="center")

    # ── Sheet 4: Live Bets (results_log.csv) ─────────────────────────────────
    if live_log:
        ls = wb.create_sheet("Live Bets")
        _title(ls, f"NHL MODEL — LIVE BET RECORD  ({len(live_log)} bets tracked)", ncols=9)
        live_hdrs = ["DATE", "GAME", "MARKET", "BET", "ODDS",
                     "EDGE%", "CONF", "RESULT", "P&L"]
        _hdr_row(ls, 2, live_hdrs)

        total_pnl = 0.0
        for ri, row in enumerate(live_log, 3):
            pnl = float(row.get("pnl", 0))
            total_pnl += pnl
            outcome = row.get("outcome", "")
            fill = GREEN if outcome == "WIN" else (RED if outcome == "LOSS" else AMBER)
            vals = [
                row.get("date", ""),
                row.get("game", ""),
                row.get("market", ""),
                row.get("bet_label", ""),
                row.get("odds", ""),
                row.get("edge_pct", ""),
                row.get("confidence", ""),
                outcome,
                f"${pnl:+.2f}",
            ]
            for ci, val in enumerate(vals, 1):
                c = ls.cell(row=ri, column=ci, value=val)
                if ci == 8:
                    c.fill = fill
                c.alignment = Alignment(horizontal="center" if ci != 2 else "left",
                                        vertical="center")
            ls.row_dimensions[ri].height = 14

        # Total row
        tr = len(live_log) + 3
        ls.merge_cells(start_row=tr, start_column=1, end_row=tr, end_column=8)
        tc = ls.cell(row=tr, column=1, value="TOTAL")
        tc.font = Font(bold=True, size=11)
        tc.alignment = Alignment(horizontal="right", vertical="center")
        uc = ls.cell(row=tr, column=9, value=f"${total_pnl:+.2f}")
        uc.font = Font(bold=True,
                       color="006100" if total_pnl >= 0 else "9C0006",
                       size=11)
        uc.alignment = Alignment(horizontal="center", vertical="center")

        for ci, w in enumerate([12, 34, 9, 28, 7, 8, 7, 9, 10], 1):
            ls.column_dimensions[get_column_letter(ci)].width = w

    # ── Sheet 5: Cumulative P&L ───────────────────────────────────────────────
    from openpyxl.chart import LineChart, Reference
    from openpyxl.chart.series import SeriesLabel

    cp = wb.create_sheet("Cumulative P&L")
    _title(cp, "NHL MODEL — CUMULATIVE UNIT P&L  (ML @ -110)", ncols=5)
    _hdr_row(cp, 2, ["#", "DATE", "GAME", "BET UNITS", "RUNNING TOTAL"])

    sorted_all = sorted(all_bets, key=lambda x: x["date"])
    running = 0.0
    for ri, b in enumerate(sorted_all, 3):
        running += b["units"]
        cp.cell(row=ri, column=1, value=ri - 2)
        cp.cell(row=ri, column=2, value=b["date"])
        cp.cell(row=ri, column=3, value=b["game"])
        cp.cell(row=ri, column=4, value=round(b["units"], 4))
        cp.cell(row=ri, column=5, value=round(running, 4))
        for col in range(1, 6):
            cp.cell(row=ri, column=col).alignment = Alignment(
                horizontal="center" if col != 3 else "left", vertical="center"
            )
        cp.row_dimensions[ri].height = 14

    # Risk stats block below data
    data_end = len(sorted_all) + 3
    risk_all = _compute_risk_metrics(sorted_all)
    cp.cell(row=data_end + 1, column=1, value="Max Drawdown").font = Font(bold=True)
    cp.cell(row=data_end + 1, column=2, value=f"{risk_all['max_drawdown']:+.2f}u")
    cp.cell(row=data_end + 2, column=1, value="Per-bet Sharpe").font = Font(bold=True)
    cp.cell(row=data_end + 2, column=2, value=f"{risk_all['sharpe']:.4f}")

    # Line chart of running total
    if len(sorted_all) >= 2:
        chart = LineChart()
        chart.title = "Cumulative Unit P&L"
        chart.style = 10
        chart.y_axis.title = "Units"
        chart.x_axis.title = "Bet #"
        chart.height = 14
        chart.width = 28
        data_ref = Reference(cp, min_col=5, min_row=2, max_row=2 + len(sorted_all))
        chart.add_data(data_ref, titles_from_data=True)
        chart.series[0].graphicalProperties.line.solidFill = "2F75B6"
        chart.series[0].graphicalProperties.line.width = 18000
        cp.add_chart(chart, f"G3")

    for ci, w in enumerate([6, 12, 32, 12, 14], 1):
        cp.column_dimensions[get_column_letter(ci)].width = w

    # ── Sheet 6: Monthly Breakdown ────────────────────────────────────────────
    mb = wb.create_sheet("Monthly Breakdown")
    _title(mb, "NHL MODEL — MONTHLY UNIT BREAKDOWN", ncols=8)
    _hdr_row(mb, 2, ["MONTH", "SEASON", "ML BETS", "ML WINS", "WIN%", "ML UNITS", "PL BETS", "PL UNITS"])

    # Aggregate by month+season
    from collections import defaultdict
    monthly: dict = defaultdict(lambda: {"ml": [], "pl": []})
    for b in results_curr["ml_bets"]:
        monthly[(b["month"], "2025-26")]["ml"].append(b)
    for b in results_prev["ml_bets"]:
        monthly[(b["month"], "2024-25")]["ml"].append(b)
    for b in results_curr["pl_bets"]:
        monthly[(b["month"], "2025-26")]["pl"].append(b)
    for b in results_prev["pl_bets"]:
        monthly[(b["month"], "2024-25")]["pl"].append(b)

    for ri, (key, data) in enumerate(sorted(monthly.items()), 3):
        month, season = key
        ml = data["ml"]
        pl = data["pl"]
        n_ml = len(ml)
        n_wins = sum(1 for b in ml if b["units"] > 0)
        ml_u = sum(b["units"] for b in ml)
        n_pl = len(pl)
        pl_u = sum(b["units"] for b in pl)
        win_pct = f"{n_wins/n_ml:.1%}" if n_ml else "—"
        vals = [month, season, n_ml, n_wins, win_pct, round(ml_u, 2), n_pl, round(pl_u, 2)]
        for ci, val in enumerate(vals, 1):
            c = mb.cell(row=ri, column=ci, value=val)
            c.alignment = Alignment(horizontal="center", vertical="center")
        # Color ML units cell
        uc = mb.cell(row=ri, column=6)
        uc.font = G_FONT if ml_u >= 0 else R_FONT
        mb.row_dimensions[ri].height = 14

    for ci, w in enumerate([10, 10, 10, 10, 8, 12, 10, 12], 1):
        mb.column_dimensions[get_column_letter(ci)].width = w

    # ── Save ──────────────────────────────────────────────────────────────────
    os.makedirs(output_dir, exist_ok=True)
    out_path = os.path.join(output_dir, "backtest_results.xlsx")
    for attempt in range(1, 6):
        try:
            wb.save(out_path)
            return out_path
        except PermissionError:
            out_path = os.path.join(output_dir, f"backtest_results_v{attempt+1}.xlsx")
    return out_path


# ---------------------------------------------------------------------------
# Live bet log loader
# ---------------------------------------------------------------------------

def _load_live_log() -> List[Dict]:
    path = os.path.join(OUTPUT_DIR, "results_log.csv")
    if not os.path.exists(path):
        return []
    with open(path, newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def main():
    args = parse_args()
    min_edge = args.min_edge

    print(Fore.CYAN + Style.BRIGHT + "\n NHL Betting Model — Historical Backtest")
    print(Fore.WHITE + f" Min edge filter: {min_edge:.1f}%   |   Simulated odds: -110")
    print(Fore.WHITE + "─" * 57 + "\n")

    # ── Load game data ────────────────────────────────────────────────────────
    print("[1/4] Loading historical game results...")
    from data.nhl_api import get_season_results
    curr_results = get_season_results(CURRENT_SEASON, GAME_TYPE_REGULAR)
    prev_results = get_season_results(PREV_SEASON, GAME_TYPE_REGULAR)
    print(f"  2025-26: {len(curr_results)} games | 2024-25: {len(prev_results)} games")

    # ── Build feature matrices ────────────────────────────────────────────────
    print("\n[2/4] Building feature matrices (this may take a minute)...")
    from features.team_stats      import build_team_stat_features
    from features.goalie_features import build_goalie_features
    from features.builder         import build_training_matrix

    team_stats   = build_team_stat_features(CURRENT_SEASON, GAME_TYPE_REGULAR)
    goalie_feats = build_goalie_features(CURRENT_SEASON, GAME_TYPE_REGULAR)

    print("  Building 2025-26 season features...")
    X_curr, y_ml_c, y_pl_h_c, y_pl_a_c, y_gh_c, y_ga_c = \
        build_training_matrix(curr_results, prev_results, team_stats, goalie_feats)
    meta_curr = _extract_metadata(curr_results, prev_results, "2025-26")

    print("  Building 2024-25 season features (out-of-sample)...")
    X_prev, y_ml_p, y_pl_h_p, y_pl_a_p, y_gh_p, y_ga_p = \
        build_training_matrix(prev_results, [], team_stats, goalie_feats)
    meta_prev = _extract_metadata(prev_results, [], "2024-25")

    print(f"  2025-26: {len(X_curr)} samples | 2024-25: {len(X_prev)} samples")

    # Verify alignment
    if len(X_curr) != len(meta_curr):
        print(Fore.YELLOW + f"  Warning: curr metadata mismatch ({len(X_curr)} vs {len(meta_curr)})")
        n = min(len(X_curr), len(meta_curr))
        X_curr = X_curr.iloc[:n]; meta_curr = meta_curr[:n]
        y_ml_c = y_ml_c.iloc[:n]; y_pl_h_c = y_pl_h_c.iloc[:n]
        y_pl_a_c = y_pl_a_c.iloc[:n]; y_gh_c = y_gh_c.iloc[:n]; y_ga_c = y_ga_c.iloc[:n]

    if len(X_prev) != len(meta_prev):
        print(Fore.YELLOW + f"  Warning: prev metadata mismatch ({len(X_prev)} vs {len(meta_prev)})")
        n = min(len(X_prev), len(meta_prev))
        X_prev = X_prev.iloc[:n]; meta_prev = meta_prev[:n]
        y_ml_p = y_ml_p.iloc[:n]; y_pl_h_p = y_pl_h_p.iloc[:n]
        y_pl_a_p = y_pl_a_p.iloc[:n]; y_gh_p = y_gh_p.iloc[:n]; y_ga_p = y_ga_p.iloc[:n]

    # ── Load trained models ───────────────────────────────────────────────────
    print("\n[3/4] Loading trained models...")
    ensemble     = joblib.load(os.path.join(SAVED_MODELS_DIR, "ensemble.joblib"))
    puckline     = joblib.load(os.path.join(SAVED_MODELS_DIR, "puckline.joblib"))
    poisson      = joblib.load(os.path.join(SAVED_MODELS_DIR, "poisson_totals.joblib"))
    train_metrics = joblib.load(os.path.join(SAVED_MODELS_DIR, "metrics.joblib"))
    print("  Models loaded from cache.")

    # ── Align feature columns to what the trained model expects ─────────────
    def _align(X: pd.DataFrame, model) -> pd.DataFrame:
        """Add any missing columns (as 0) so the model's feature set is satisfied."""
        expected = getattr(model.logistic, "feature_names", None) or \
                   getattr(model, "feature_names_in_", None)
        if expected is None:
            return X
        missing = [c for c in expected if c not in X.columns]
        if missing:
            for col in missing:
                X = X.copy()
                X[col] = 0.0
        return X

    X_curr = _align(X_curr, ensemble)
    X_prev = _align(X_prev, ensemble)

    # ── Run predictions ───────────────────────────────────────────────────────
    print("\n[4/4] Running predictions on all historical games...")
    probs_curr, _ = ensemble.predict_proba(X_curr)
    probs_prev, _ = ensemble.predict_proba(X_prev)

    pl_h_curr = puckline.predict_proba_home_minus1_5(X_curr)
    pl_a_curr = puckline.predict_proba_away_plus1_5(X_curr)
    pl_h_prev = puckline.predict_proba_home_minus1_5(X_prev)
    pl_a_prev = puckline.predict_proba_away_plus1_5(X_prev)

    mu_h_curr, mu_a_curr = poisson.predict_goals(X_curr)
    mu_h_prev, mu_a_prev = poisson.predict_goals(X_prev)
    print("  Done.\n")

    # Propagate season label to bet log
    for m in meta_curr:
        m["season"] = "2025-26"
    for m in meta_prev:
        m["season"] = "2024-25"

    # ── Print season reports ──────────────────────────────────────────────────
    print(Fore.WHITE + "=" * 59)
    print(Fore.CYAN + Style.BRIGHT + "  NHL BETTING MODEL — BACKTEST RESULTS")
    print(Fore.WHITE + "=" * 59)

    results_prev = _season_block(
        label          = "2024-25 SEASON  (out-of-sample)",
        probs          = probs_prev,
        actuals        = y_ml_p.values,
        pl_h_probs     = pl_h_prev,
        y_pl_h         = y_pl_h_p.values,
        pl_a_probs     = pl_a_prev,
        y_pl_a         = y_pl_a_p.values,
        mu_h           = mu_h_prev,
        mu_a           = mu_a_prev,
        y_gh           = y_gh_p.values,
        y_ga           = y_ga_p.values,
        meta           = meta_prev,
        train_val_metrics = None,
        min_edge       = min_edge,
    )

    results_curr = _season_block(
        label          = "2025-26 SEASON  (YTD)",
        probs          = probs_curr,
        actuals        = y_ml_c.values,
        pl_h_probs     = pl_h_curr,
        y_pl_h         = y_pl_h_c.values,
        pl_a_probs     = pl_a_curr,
        y_pl_a         = y_pl_a_c.values,
        mu_h           = mu_h_curr,
        mu_a           = mu_a_curr,
        y_gh           = y_gh_c.values,
        y_ga           = y_ga_c.values,
        meta           = meta_curr,
        train_val_metrics = train_metrics,
        min_edge       = min_edge,
    )

    # ── Combined summary ──────────────────────────────────────────────────────
    all_bets  = results_prev["ml_bets"] + results_curr["ml_bets"]
    all_units = sum(b["units"] for b in all_bets)
    all_pl    = results_prev["pl_bets"] + results_curr["pl_bets"]
    all_pl_u  = sum(b["units"] for b in all_pl)

    _header("COMBINED — BOTH SEASONS", Fore.MAGENTA)
    _row("Total ML bets simulated",  str(len(all_bets)))
    _row("Total ML units",           _fmt_u(all_units),
         color=Fore.GREEN if all_units >= 0 else Fore.RED, bold=True)
    _row("Total PL bets simulated",  str(len(all_pl)))
    _row("Total PL units",           _fmt_u(all_pl_u),
         color=Fore.GREEN if all_pl_u >= 0 else Fore.RED)

    # ── Live bet tracking (results_log.csv) ───────────────────────────────────
    live_log = _load_live_log()
    if live_log:
        _header("LIVE BET TRACKING  (results_log.csv)", Fore.YELLOW)
        n_live = len(live_log)
        live_pnl = sum(float(r.get("pnl", 0)) for r in live_log)
        live_units = live_pnl / 10  # $10/unit
        wins  = sum(1 for r in live_log if r.get("outcome") == "WIN")
        losses = sum(1 for r in live_log if r.get("outcome") == "LOSS")
        _row("Bets tracked", str(n_live))
        _row("W / L",        f"{wins} / {losses}  ({wins/n_live:.1%} win rate)")
        _row("Total P&L",    f"${live_pnl:+.2f}  ({live_units:+.2f}u at $10/unit)",
             color=Fore.GREEN if live_pnl >= 0 else Fore.RED, bold=True)
    else:
        print(Fore.WHITE + "\n  (No live bet records found — run --track to start logging)\n")

    # ── Export ────────────────────────────────────────────────────────────────
    if args.export == "excel":
        print(Fore.WHITE + "\n  Exporting to Excel...")
        path = _export_excel(results_curr, results_prev, live_log, OUTPUT_DIR)
        print(Fore.GREEN + f"  Saved: {path}\n")
    else:
        print(Fore.WHITE + f"\n  Tip: add --export excel to save full report to Excel.\n")


if __name__ == "__main__":
    main()
