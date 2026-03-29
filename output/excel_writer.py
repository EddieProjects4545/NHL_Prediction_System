"""
Excel output builder — produces a color-coded daily predictions workbook.

Layout: top-down vertical blocks (one 6-row block per game, 10 columns wide).
Right-side summary panel (cols L–S) shows HIGH and MEDIUM confidence best bets.
No horizontal scrolling required for the main grid.
"""
import logging
import os
import re
from typing import Dict, List, Optional

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from config import OUTPUT_DIR

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Style constants
# ---------------------------------------------------------------------------

FILLS = {
    "HIGH":   PatternFill(fill_type="solid", fgColor="C6EFCE"),   # soft green
    "MEDIUM": PatternFill(fill_type="solid", fgColor="FFEB9C"),   # soft yellow
    "LOW":    PatternFill(fill_type="solid", fgColor="FCE4D6"),   # soft peach
    "NONE":   PatternFill(fill_type="solid", fgColor="F2F2F2"),   # light grey (no rec)
}

SECTION_FILLS = {
    "game":    PatternFill(fill_type="solid", fgColor="1F4E79"),  # dark navy
    "ml":      PatternFill(fill_type="solid", fgColor="375623"),  # dark green
    "pl":      PatternFill(fill_type="solid", fgColor="44546A"),  # dark slate
    "ou":      PatternFill(fill_type="solid", fgColor="7030A0"),  # purple
}

HEADER_FILL    = PatternFill(fill_type="solid", fgColor="2F75B6")
HEADER_FONT    = Font(bold=True, color="FFFFFF")
SECTION_FONT   = Font(bold=True, color="FFFFFF", size=11)

PRED_ROW_FILL    = PatternFill(fill_type="solid", fgColor="2E4057")  # dark slate
GAME_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
PRED_ROW_FONT    = Font(italic=True, color="FFFFFF", size=10)
SUBHEADER_FONT   = Font(bold=True, color="FFFFFF", size=9)

WIN_FILL   = PatternFill(fill_type="solid", fgColor="C6EFCE")
LOSS_FILL  = PatternFill(fill_type="solid", fgColor="FFCCCC")
PUSH_FILL  = PatternFill(fill_type="solid", fgColor="FFEB9C")
WIN_FONT   = Font(bold=True, color="006100")
LOSS_FONT  = Font(bold=True, color="9C0006")
PUSH_FONT  = Font(bold=True, color="7D6608")

# ---------------------------------------------------------------------------
# Layout constants
# ---------------------------------------------------------------------------

ROWS_PER_GAME      = 6
DATA_START_ROW     = 3   # row 1 = title, row 2 = blank spacer
GAME_HEADER_OFFSET = 0
PRED_ROW_OFFSET    = 1
SUBHEADER_OFFSET   = 2
ML_ROW_OFFSET      = 3
PL_ROW_OFFSET      = 4
OU_ROW_OFFSET      = 5

TOTAL_COLS   = 10
COL_BET_TYPE = 1
COL_PICK     = 2
COL_CONF     = 3
COL_ODDS     = 4
COL_MODEL    = 5
COL_MARKET   = 6
COL_EDGE     = 7
COL_EV       = 8
COL_UNITS_COL = 9
COL_RESULT   = 10

_SUBHEADER_LABELS = [
    "BET TYPE", "PICK", "CONF", "ODDS",
    "MODEL%", "MARKET%", "EDGE", "EV%", "SIZE", "RESULT",
]
_COL_WIDTHS = [16, 26, 9, 9, 9, 9, 9, 9, 9, 10]

# Right-side summary panel
_SUMMARY_START_COL = 12   # column L
_SUMMARY_HEADERS   = ["GAME", "TYPE", "PICK", "ODDS", "EDGE", "SIZE", "RESULT", "UNITS"]
_SUMMARY_WIDTHS    = [26, 8, 24, 8, 8, 8, 9, 9]

SUMMARY_HIGH_FILL = PatternFill(fill_type="solid", fgColor="1A6B2A")   # dark green
SUMMARY_MED_FILL  = PatternFill(fill_type="solid", fgColor="7D5A00")   # dark amber
SUMMARY_ROW_FONT  = Font(color="000000", size=10)


# ---------------------------------------------------------------------------
# Confidence tier
# ---------------------------------------------------------------------------

def _conf_tier(model_prob: float, edge_pct: float) -> str:
    """
    Derive display tier from model probability and edge.

    HIGH   : model win prob >= 60%  OR  edge >= 8%
    MEDIUM : edge 4% – 7.9%
    LOW    : edge 1% – 3.9%
    """
    if model_prob >= 0.60 or edge_pct >= 8.0:
        return "HIGH"
    if edge_pct >= 4.0:
        return "MEDIUM"
    return "LOW"


# ---------------------------------------------------------------------------
# Public: build workbook
# ---------------------------------------------------------------------------

def build_daily_workbook(date_str: str, game_results: List[dict]) -> openpyxl.Workbook:
    """
    Build a top-down predictions workbook. Each game occupies a 6-row block:
        Row +0  Game header (merged A:J) — dark navy
        Row +1  Prediction summary (merged A:J) — dark slate
        Row +2  Sub-header labels
        Row +3  Moneyline row
        Row +4  Puck Line row
        Row +5  Over/Under row
        [spacer]
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = date_str

    # Row 1: sheet title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=TOTAL_COLS)
    title_cell = ws.cell(row=1, column=1,
                         value=f"NHL Predictions — {date_str}  ({len(game_results)} games)")
    title_cell.fill      = SECTION_FILLS["game"]
    title_cell.font      = Font(bold=True, color="FFFFFF", size=13)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    # Row 2: blank spacer
    ws.row_dimensions[2].height = 6

    _set_column_widths(ws)
    ws.freeze_panes = "A3"

    for game_idx, game in enumerate(game_results):
        base_row = DATA_START_ROW + game_idx * ROWS_PER_GAME
        _write_game_block(ws, base_row, game)

    _write_confidence_summary(ws, game_results)

    build_record_sheet(wb, OUTPUT_DIR, skip_date=date_str)

    return wb


# ---------------------------------------------------------------------------
# Public: save workbook
# ---------------------------------------------------------------------------

def save_workbook(wb: openpyxl.Workbook, date_str: str,
                  output_dir: str = None) -> str:
    """Save workbook to disk. Appends _v2, _v3 etc. if file is locked."""
    out_dir = output_dir or OUTPUT_DIR
    os.makedirs(out_dir, exist_ok=True)

    path = os.path.join(out_dir, f"predictions_{date_str}.xlsx")

    for attempt in range(1, 10):
        try:
            wb.save(path)
            logger.info("Saved predictions to %s", path)
            return path
        except PermissionError:
            suffix = f"_v{attempt + 1}"
            path = os.path.join(out_dir, f"predictions_{date_str}{suffix}.xlsx")
            logger.warning("File locked — trying %s", path)

    logger.error("Could not save workbook after multiple attempts.")
    return path


# ---------------------------------------------------------------------------
# Public: write results back after games complete
# ---------------------------------------------------------------------------

def write_results_to_workbook(date_str: str,
                               completed_games: List[dict],
                               output_dir: str = None) -> None:
    """
    Open a predictions workbook and write actual scores + WIN/LOSS into the
    RESULT column of each ML / PL / OU row.

    completed_games: list of dicts with keys:
        home_name (abbrev), away_name (abbrev), home_score (int), away_score (int)
    """
    out_dir = output_dir or OUTPUT_DIR
    path = os.path.join(out_dir, f"predictions_{date_str}.xlsx")

    if not os.path.exists(path):
        logger.warning("No predictions file found for %s — skipping result write-back.", date_str)
        return

    try:
        wb = openpyxl.load_workbook(path)
    except Exception as e:
        logger.error("Could not open %s: %s", path, e)
        return

    ws = wb.active

    # Build lookup: norm(game_label) -> {home, away scores}
    results_lookup: Dict[str, dict] = {}
    for g in completed_games:
        home = g.get("home_name", "")
        away = g.get("away_name", "")
        label = f"{away} @ {home}"
        try:
            results_lookup[_norm(label)] = {
                "home": int(g["home_score"]),
                "away": int(g["away_score"]),
            }
        except (TypeError, ValueError):
            pass

    # Scan for "Moneyline" anchor rows to locate each game block
    for row_idx in range(1, ws.max_row + 1):
        if ws.cell(row=row_idx, column=COL_BET_TYPE).value != "Moneyline":
            continue

        base_row   = row_idx - ML_ROW_OFFSET
        header_val = str(ws.cell(row=base_row + GAME_HEADER_OFFSET, column=1).value or "")
        game_label = _extract_game_label(header_val)

        if _norm(game_label) not in results_lookup:
            continue

        res   = results_lookup[_norm(game_label)]
        hs    = res["home"]
        as_   = res["away"]
        total = hs + as_

        # Patch prediction row: replace "Actual: ___ – ___" placeholder
        pred_row  = base_row + PRED_ROW_OFFSET
        pred_cell = ws.cell(row=pred_row, column=1)
        new_text  = str(pred_cell.value or "").replace(
            "Actual: ___ \u2013 ___", f"Actual: {hs} \u2013 {as_}"
        )
        pred_cell.value = new_text
        pred_cell.font  = PRED_ROW_FONT

        # Parse home/away abbreviations from game label ("FLA @ CAR")
        label_parts = game_label.split("@")
        away_abbr = label_parts[0].strip().upper() if len(label_parts) == 2 else ""
        home_abbr = label_parts[1].strip().upper() if len(label_parts) == 2 else ""

        # Evaluate results
        ml_pick = str(ws.cell(row=base_row + ML_ROW_OFFSET, column=COL_PICK).value or "")
        pl_pick = str(ws.cell(row=base_row + PL_ROW_OFFSET, column=COL_PICK).value or "")
        ou_pick = str(ws.cell(row=base_row + OU_ROW_OFFSET, column=COL_PICK).value or "")

        ml_result, ml_fill = _eval_ml_pick(ml_pick, home_abbr, away_abbr, hs, as_)
        pl_result, pl_fill = _eval_pl_pick(pl_pick, home_abbr, away_abbr, hs, as_)
        ou_result, ou_fill = _eval_ou(ou_pick, total)

        for result_row, result, rfill in [
            (base_row + ML_ROW_OFFSET, ml_result, ml_fill),
            (base_row + PL_ROW_OFFSET, pl_result, pl_fill),
            (base_row + OU_ROW_OFFSET, ou_result, ou_fill),
        ]:
            if not result:
                continue
            cell = ws.cell(row=result_row, column=COL_RESULT, value=result)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if rfill:
                cell.fill = rfill
                cell.font = WIN_FONT if rfill == WIN_FILL else (
                    LOSS_FONT if rfill == LOSS_FILL else PUSH_FONT
                )

    # Update right-side summary panel with RESULT + UNITS
    _BET_LABELS = {"Moneyline": "ML", "Puck Line": "PL", "Over/Under": "OU"}
    _OFFSETS    = {"Moneyline": ML_ROW_OFFSET, "Puck Line": PL_ROW_OFFSET, "Over/Under": OU_ROW_OFFSET}

    # Build lookup: norm(game_label) -> {type -> (result, odds)}
    summary_results: Dict[str, Dict[str, tuple]] = {}
    for row_idx in range(1, ws.max_row + 1):
        bet_label = ws.cell(row=row_idx, column=COL_BET_TYPE).value
        if bet_label not in _BET_LABELS:
            continue
        offset    = _OFFSETS[bet_label]
        hdr_val   = str(ws.cell(row=row_idx - offset, column=1).value or "")
        game_lbl  = _extract_game_label(hdr_val)
        result_v  = str(ws.cell(row=row_idx, column=COL_RESULT).value or "")
        odds_v    = str(ws.cell(row=row_idx, column=COL_ODDS).value or "")
        key       = _norm(game_lbl)
        if key not in summary_results:
            summary_results[key] = {}
        summary_results[key][_BET_LABELS[bet_label]] = (result_v, odds_v)

    _COL_S_GAME   = _SUMMARY_START_COL          # L
    _COL_S_TYPE   = _SUMMARY_START_COL + 1      # M
    _COL_S_RESULT = _SUMMARY_START_COL + 6      # R
    _COL_S_UNITS  = _SUMMARY_START_COL + 7      # S

    for row_idx in range(1, ws.max_row + 1):
        game_val = ws.cell(row=row_idx, column=_COL_S_GAME).value
        type_val = str(ws.cell(row=row_idx, column=_COL_S_TYPE).value or "").upper()
        if not game_val or type_val not in ("ML", "PL", "OU"):
            continue
        key = _norm(str(game_val))
        if key not in summary_results or type_val not in summary_results[key]:
            continue
        result_v, odds_v = summary_results[key][type_val]
        units = _calc_units(result_v, odds_v)

        r_cell = ws.cell(row=row_idx, column=_COL_S_RESULT, value=result_v)
        r_cell.alignment = Alignment(horizontal="center", vertical="center")
        r_cell.font = WIN_FONT if result_v == "WIN" else (
            LOSS_FONT if result_v == "LOSS" else PUSH_FONT
        )
        r_cell.fill = WIN_FILL if result_v == "WIN" else (
            LOSS_FILL if result_v == "LOSS" else PUSH_FILL
        )

        u_cell = ws.cell(row=row_idx, column=_COL_S_UNITS, value=_fmt_units(units))
        u_cell.alignment = Alignment(horizontal="center", vertical="center")
        u_cell.font = WIN_FONT if units > 0 else (
            LOSS_FONT if units < 0 else PUSH_FONT
        )

    # Refresh record sheet
    build_record_sheet(wb, out_dir)

    # Lock-safe save
    for attempt in range(1, 10):
        try:
            wb.save(path)
            logger.info("Results written to %s", path)
            break
        except PermissionError:
            alt = os.path.join(out_dir, f"predictions_{date_str}_v{attempt + 1}.xlsx")
            logger.warning("File locked — saving to %s", alt)
            path = alt
    else:
        logger.error("Could not save results workbook.")
        return

    # Refresh season record
    build_season_record(out_dir)


# ---------------------------------------------------------------------------
# Public: record sheet (embedded in each daily workbook)
# ---------------------------------------------------------------------------

def build_record_sheet(wb: openpyxl.Workbook, outputs_dir: str = None,
                       skip_date: str = None) -> None:
    """
    Scan all predictions_*.xlsx files in outputs_dir, tally WIN/LOSS/PUSH
    from each bet row, and write/replace a 'Model Record' sheet in wb.
    skip_date: date string (e.g. '2026-03-29') — skip that file to avoid
               Windows file-handle conflicts when the workbook is being built.
    """
    if "Model Record" in wb.sheetnames:
        del wb["Model Record"]

    rs = wb.create_sheet("Model Record")

    tally = {
        "all":  {"ML": [0, 0, 0], "PL": [0, 0, 0], "OU": [0, 0, 0]},
        "high": {"ML": [0, 0, 0], "PL": [0, 0, 0], "OU": [0, 0, 0]},
    }

    bet_type_map = {
        "Moneyline":  "ML",
        "Puck Line":  "PL",
        "Over/Under": "OU",
    }

    out_dir = outputs_dir or OUTPUT_DIR
    skip_fname = f"predictions_{skip_date}.xlsx" if skip_date else None
    xlsx_files = []
    if os.path.isdir(out_dir):
        for f in sorted(os.listdir(out_dir)):
            if f.startswith("predictions_") and f.endswith(".xlsx"):
                if f == skip_fname:
                    continue
                xlsx_files.append(os.path.join(out_dir, f))

    scanned = 0
    for fpath in xlsx_files:
        file_wb = None
        try:
            file_wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
            file_ws = file_wb.active

            for row in file_ws.iter_rows(min_col=1, max_col=TOTAL_COLS):
                bet_label = row[COL_BET_TYPE - 1].value
                if bet_label not in bet_type_map:
                    continue
                bet_key = bet_type_map[bet_label]
                conf    = str(row[COL_CONF   - 1].value or "").upper()
                result  = str(row[COL_RESULT - 1].value or "").upper()

                if result == "WIN":
                    tally["all"][bet_key][0] += 1
                    if conf == "HIGH":
                        tally["high"][bet_key][0] += 1
                elif result == "LOSS":
                    tally["all"][bet_key][1] += 1
                    if conf == "HIGH":
                        tally["high"][bet_key][1] += 1
                elif result == "PUSH":
                    tally["all"][bet_key][2] += 1
                    if conf == "HIGH":
                        tally["high"][bet_key][2] += 1

            scanned += 1
        except Exception as e:
            logger.warning("Could not scan %s: %s", fpath, e)
        finally:
            if file_wb is not None:
                file_wb.close()

    _write_record_sheet(rs, tally, scanned)


# ---------------------------------------------------------------------------
# Public: season record (standalone season_record.xlsx)
# ---------------------------------------------------------------------------

def build_season_record(outputs_dir: str = None) -> None:
    """
    Build/rebuild season_record.xlsx with two sheets:
      - Pick Log : one row per HIGH/MEDIUM pick with result + units P/L
      - Tally    : W/L/Win%/Units broken out by ALL / HIGH / MEDIUM and ML/PL/OU
    """
    out_dir    = outputs_dir or OUTPUT_DIR
    record_path = os.path.join(out_dir, "season_record.xlsx")

    bet_type_map = {"Moneyline": "ML", "Puck Line": "PL", "Over/Under": "OU"}
    picks: List[dict] = []

    xlsx_files = []
    if os.path.isdir(out_dir):
        for f in sorted(os.listdir(out_dir)):
            if (f.startswith("predictions_") and f.endswith(".xlsx")
                    and "season_record" not in f):
                xlsx_files.append(os.path.join(out_dir, f))

    for fpath in xlsx_files:
        try:
            date_str = os.path.basename(fpath).replace("predictions_", "")[:10]
        except Exception:
            date_str = "unknown"
        try:
            file_wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
            file_ws = file_wb.active
        except Exception as e:
            logger.warning("Could not open %s: %s", fpath, e)
            continue

        current_game = ""
        for row in file_ws.iter_rows(min_col=1, max_col=TOTAL_COLS):
            col1 = str(row[0].value or "")
            if " @ " in col1 and row[COL_BET_TYPE - 1].value not in bet_type_map:
                current_game = _extract_game_label(col1)

            bet_label = row[COL_BET_TYPE - 1].value
            if bet_label not in bet_type_map:
                continue

            conf = str(row[COL_CONF - 1].value or "").upper()
            if conf not in ("HIGH", "MEDIUM"):
                continue

            result = str(row[COL_RESULT - 1].value or "")
            odds   = str(row[COL_ODDS   - 1].value or "")
            pick   = str(row[COL_PICK   - 1].value or "")
            edge   = str(row[COL_EDGE   - 1].value or "")
            units  = _calc_units(result, odds)

            picks.append({
                "date":   date_str,
                "game":   current_game,
                "type":   bet_type_map[bet_label],
                "conf":   conf,
                "pick":   pick,
                "odds":   odds,
                "edge":   edge,
                "result": result,
                "units":  units,
            })
        file_wb.close()

    # Build tally
    tally = {
        tier: {"ML": [0, 0, 0, 0.0], "PL": [0, 0, 0, 0.0], "OU": [0, 0, 0, 0.0]}
        for tier in ("all", "high", "medium")
    }
    for p in picks:
        t = p["type"]
        u = p["units"]
        r = p["result"].upper()
        w  = 1 if r == "WIN"  else 0
        l  = 1 if r == "LOSS" else 0
        ps = 1 if r == "PUSH" else 0
        for tier in ("all", p["conf"].lower()):
            tally[tier][t][0] += w
            tally[tier][t][1] += l
            tally[tier][t][2] += ps
            tally[tier][t][3] += u

    wb = openpyxl.Workbook()

    # ── Sheet 1: Pick Log ──────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Pick Log"

    log_headers = ["DATE", "GAME", "TYPE", "CONF", "PICK", "ODDS", "EDGE", "RESULT", "UNITS"]
    log_widths   = [12, 32, 8, 10, 26, 8, 9, 9, 10]

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(log_headers))
    t = ws.cell(row=1, column=1, value="NHL SEASON PICK LOG  (HIGH & MEDIUM confidence)")
    t.fill      = SECTION_FILLS["game"]
    t.font      = Font(bold=True, color="FFFFFF", size=13)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    for ci, lbl in enumerate(log_headers, 1):
        c = ws.cell(row=2, column=ci, value=lbl)
        c.fill      = HEADER_FILL
        c.font      = HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 16

    for ri, p in enumerate(picks, 3):
        row_fill = (PatternFill(fill_type="solid", fgColor="E2EFDA") if p["conf"] == "HIGH"
                    else PatternFill(fill_type="solid", fgColor="FFF2CC"))
        vals = [p["date"], p["game"], p["type"], p["conf"],
                p["pick"], p["odds"], p["edge"], p["result"],
                _fmt_units(p["units"])]
        for ci, val in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.fill      = row_fill
            cell.alignment = Alignment(
                horizontal="left" if ci == 2 else "center",
                vertical="center"
            )
        r_cell = ws.cell(row=ri, column=8)
        u_cell = ws.cell(row=ri, column=9)
        r = p["result"].upper()
        if r == "WIN":
            r_cell.font = WIN_FONT
            u_cell.font = WIN_FONT
        elif r == "LOSS":
            r_cell.font = LOSS_FONT
            u_cell.font = LOSS_FONT
        ws.row_dimensions[ri].height = 15

    if picks:
        total_units = sum(p["units"] for p in picks)
        total_row   = len(picks) + 3
        ws.merge_cells(start_row=total_row, start_column=1,
                        end_row=total_row, end_column=8)
        tc = ws.cell(row=total_row, column=1, value="SEASON TOTAL")
        tc.font      = Font(bold=True, size=11)
        tc.alignment = Alignment(horizontal="right", vertical="center")
        uc = ws.cell(row=total_row, column=9, value=_fmt_units(total_units))
        uc.font      = Font(
            bold=True,
            color="006100" if total_units >= 0 else "9C0006",
            size=11,
        )
        uc.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[total_row].height = 18

    for ci, w in enumerate(log_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # ── Sheet 2: Tally ─────────────────────────────────────────────────────────
    ts = wb.create_sheet("Tally")

    def _win_pct(w, l):
        return f"{w / (w + l) * 100:.1f}%" if (w + l) > 0 else "\u2014"

    def _write_tally_section(start_row: int, label: str, tier_key: str) -> None:
        ts.merge_cells(start_row=start_row, start_column=1,
                        end_row=start_row, end_column=8)
        sh = ts.cell(row=start_row, column=1, value=label)
        sh.fill      = SECTION_FILLS["game"]
        sh.font      = SECTION_FONT
        sh.alignment = Alignment(horizontal="center", vertical="center")
        ts.row_dimensions[start_row].height = 22

        hdrs = ["TYPE", "W", "L", "PUSH", "TOTAL", "WIN%", "UNITS", "$ P/L ($10/u)"]
        for ci, h in enumerate(hdrs, 1):
            c = ts.cell(row=start_row + 1, column=ci, value=h)
            c.fill      = HEADER_FILL
            c.font      = HEADER_FONT
            c.alignment = Alignment(horizontal="center", vertical="center")
        ts.row_dimensions[start_row + 1].height = 16

        total_w = total_l = total_p = 0
        total_u = 0.0
        for offset, (bt, label2) in enumerate(
            [("ML", "Moneyline"), ("PL", "Puck Line"), ("OU", "Over/Under")], 2
        ):
            w, l, p, u = tally[tier_key][bt]
            total_w += w
            total_l += l
            total_p += p
            total_u += u
            row_data = [label2, w, l, p, w + l + p,
                        _win_pct(w, l), _fmt_units(u), f"${u * 10:+.2f}"]
            for ci, val in enumerate(row_data, 1):
                c = ts.cell(row=start_row + offset, column=ci, value=val)
                c.alignment = Alignment(horizontal="center", vertical="center")
            ts.row_dimensions[start_row + offset].height = 15

        cmb_row = start_row + 5
        cmb = [" COMBINED", total_w, total_l, total_p,
               total_w + total_l + total_p, _win_pct(total_w, total_l),
               _fmt_units(total_u), f"${total_u * 10:+.2f}"]
        for ci, val in enumerate(cmb, 1):
            c = ts.cell(row=cmb_row, column=ci, value=val)
            c.font      = Font(bold=True)
            c.alignment = Alignment(horizontal="center", vertical="center")
        ts.row_dimensions[cmb_row].height = 16

    ts.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    title = ts.cell(row=1, column=1, value="NHL MODEL — SEASON RECORD")
    title.fill      = SECTION_FILLS["game"]
    title.font      = Font(bold=True, color="FFFFFF", size=14)
    title.alignment = Alignment(horizontal="center", vertical="center")
    ts.row_dimensions[1].height = 28

    _write_tally_section(3,  "ALL PICKS (HIGH + MEDIUM)", "all")
    _write_tally_section(10, "HIGH CONFIDENCE",            "high")
    _write_tally_section(17, "MEDIUM CONFIDENCE",          "medium")

    for ci, w in enumerate([16, 8, 8, 8, 8, 9, 10, 14], 1):
        ts.column_dimensions[get_column_letter(ci)].width = w

    for attempt in range(1, 6):
        try:
            wb.save(record_path)
            logger.info("Season record saved to %s", record_path)
            return
        except PermissionError:
            record_path = os.path.join(out_dir, f"season_record_v{attempt + 1}.xlsx")

    logger.error("Could not save season_record.xlsx")


# ---------------------------------------------------------------------------
# Internal: game block writer
# ---------------------------------------------------------------------------

def _write_game_block(ws, base_row: int, game: dict) -> None:
    ml_rec = game.get("ml")
    pl_rec = game.get("pl")
    ou_rec = game.get("ou")

    home          = game.get("home_team", "")
    away          = game.get("away_team", "")
    home_goalie   = game.get("home_goalie", "TBD")
    away_goalie   = game.get("away_goalie", "TBD")
    home_sv       = game.get("home_goalie_sv", 0.0)
    away_sv       = game.get("away_goalie_sv", 0.0)
    pred_home     = game.get("pred_home")
    pred_away     = game.get("pred_away")
    pred_total    = game.get("pred_total")
    home_win_prob = game.get("home_win_prob")

    hg_str = f"{home_goalie} ({home_sv:.3f})" if home_sv else home_goalie
    ag_str = f"{away_goalie} ({away_sv:.3f})" if away_sv else away_goalie

    # Row +0: game header
    hdr_row  = base_row + GAME_HEADER_OFFSET
    hdr_text = (f"{game.get('game_label', f'{away} @ {home}')}"
                f"  |  Home G: {hg_str}"
                f"  |  Away G: {ag_str}")
    ws.merge_cells(start_row=hdr_row, start_column=1, end_row=hdr_row, end_column=TOTAL_COLS)
    cell = ws.cell(row=hdr_row, column=1, value=hdr_text)
    cell.fill      = SECTION_FILLS["game"]
    cell.font      = GAME_HEADER_FONT
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[hdr_row].height = 22

    # Row +1: prediction summary
    pred_row  = base_row + PRED_ROW_OFFSET
    pred_text = (
        f"Predicted:  {home} {_fmt_score(pred_home)}  \u2013  "
        f"{away} {_fmt_score(pred_away)}  |  "
        f"Total: {_fmt_score(pred_total)}  |  "
        f"Home Win: {_fmt_pct(home_win_prob)}  |  "
        f"Actual: ___ \u2013 ___"
    )
    ws.merge_cells(start_row=pred_row, start_column=1, end_row=pred_row, end_column=TOTAL_COLS)
    pred_cell = ws.cell(row=pred_row, column=1, value=pred_text)
    pred_cell.fill      = PRED_ROW_FILL
    pred_cell.font      = PRED_ROW_FONT
    pred_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[pred_row].height = 18

    # Row +2: sub-header
    sub_row = base_row + SUBHEADER_OFFSET
    for col_i, label in enumerate(_SUBHEADER_LABELS, start=1):
        cell = ws.cell(row=sub_row, column=col_i, value=label)
        cell.fill      = HEADER_FILL
        cell.font      = SUBHEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[sub_row].height = 16

    # Rows +3 / +4 / +5: ML, PL, OU
    _write_bet_row(ws, base_row + ML_ROW_OFFSET, "Moneyline",  ml_rec)
    _write_bet_row(ws, base_row + PL_ROW_OFFSET, "Puck Line",  pl_rec)
    _write_bet_row(ws, base_row + OU_ROW_OFFSET, "Over/Under", ou_rec)

    # Spacer
    ws.row_dimensions[base_row + ROWS_PER_GAME].height = 8


def _write_bet_row(ws, row_idx: int, label: str, rec: Optional[dict]) -> None:
    """Write a single ML / PL / OU data row. rec=None → empty data cells."""
    if rec:
        tier      = _conf_tier(rec.get("model_prob", 0.0), rec.get("edge_pct", 0.0))
        kelly_str = "1u"
        values = [
            label,
            rec.get("pick", ""),
            tier,
            _fmt_odds(rec.get("odds")),
            _fmt_pct(rec.get("model_prob")),
            _fmt_pct(rec.get("market_prob")),
            _fmt_edge(rec.get("edge_pct")),
            _fmt_edge(rec.get("ev_pct")),
            kelly_str,
            "",  # RESULT — filled by write_results_to_workbook
        ]
        row_fill = FILLS.get(tier, FILLS["LOW"])
    else:
        values   = [label, "", "", "", "", "", "", "", "", ""]
        row_fill = FILLS["NONE"]

    for col_i, val in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=col_i, value=val)
        cell.fill      = row_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if col_i == COL_PICK:
            cell.alignment = Alignment(horizontal="left", vertical="center")

    ws.row_dimensions[row_idx].height = 16


# ---------------------------------------------------------------------------
# Internal: right-side confidence summary panel
# ---------------------------------------------------------------------------

def _write_confidence_summary(ws, game_results: List[dict]) -> None:
    """
    Write HIGH / MEDIUM confidence best-bets panel starting at column L.
    Sorted by edge_pct descending within each tier.
    """
    col = _SUMMARY_START_COL

    for offset, width in enumerate(_SUMMARY_WIDTHS):
        ws.column_dimensions[get_column_letter(col + offset)].width = width

    high_picks: List[tuple] = []
    med_picks:  List[tuple] = []

    for game in game_results:
        label = game.get("game_label", "")
        for key, type_label in [("ml", "ML"), ("pl", "PL"), ("ou", "OU")]:
            rec = game.get(key)
            if rec is None:
                continue
            tier = _conf_tier(rec.get("model_prob", 0.0), rec.get("edge_pct", 0.0))
            if tier == "HIGH":
                high_picks.append((label, type_label, rec))
            elif tier == "MEDIUM":
                med_picks.append((label, type_label, rec))

    # Sort by edge descending within each tier
    high_picks.sort(key=lambda x: x[2].get("edge_pct", 0), reverse=True)
    med_picks.sort(key=lambda x:  x[2].get("edge_pct", 0), reverse=True)

    current_row = 1

    def _write_section_header(row: int, title: str, fill) -> None:
        ws.merge_cells(start_row=row, start_column=col,
                        end_row=row, end_column=col + len(_SUMMARY_HEADERS) - 1)
        cell = ws.cell(row=row, column=col, value=title)
        cell.fill      = fill
        cell.font      = Font(bold=True, color="FFFFFF", size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = 22

    def _write_col_headers(row: int) -> None:
        for offset, lbl in enumerate(_SUMMARY_HEADERS):
            cell = ws.cell(row=row, column=col + offset, value=lbl)
            cell.fill      = HEADER_FILL
            cell.font      = SUBHEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = 16

    def _write_pick_row(row: int, game_label: str, type_label: str,
                        rec: dict, row_fill) -> None:
        kelly_str = "1u"
        values = [
            game_label,
            type_label,
            rec.get("pick", ""),
            _fmt_odds(rec.get("odds")),
            _fmt_edge(rec.get("edge_pct")),
            kelly_str,
            "",   # RESULT
            "",   # UNITS
        ]
        for offset, val in enumerate(values):
            cell = ws.cell(row=row, column=col + offset, value=val)
            cell.fill      = row_fill
            cell.font      = SUMMARY_ROW_FONT
            cell.alignment = Alignment(
                horizontal="left" if offset == 0 else "center",
                vertical="center",
            )
        ws.row_dimensions[row].height = 16

    # HIGH section
    _write_section_header(current_row,
                           f"HIGH CONFIDENCE PICKS  ({len(high_picks)})",
                           SUMMARY_HIGH_FILL)
    current_row += 1
    if high_picks:
        _write_col_headers(current_row)
        current_row += 1
        high_fill = PatternFill(fill_type="solid", fgColor="E2EFDA")
        for game_label, type_label, rec in high_picks:
            _write_pick_row(current_row, game_label, type_label, rec, high_fill)
            current_row += 1
    else:
        ws.merge_cells(start_row=current_row, start_column=col,
                        end_row=current_row, end_column=col + len(_SUMMARY_HEADERS) - 1)
        cell = ws.cell(row=current_row, column=col,
                        value="No HIGH confidence picks today")
        cell.font      = Font(italic=True, color="595959", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        current_row += 1

    # Spacer row
    ws.row_dimensions[current_row].height = 10
    current_row += 1

    # MEDIUM section
    _write_section_header(current_row,
                           f"MEDIUM CONFIDENCE PICKS  ({len(med_picks)})",
                           SUMMARY_MED_FILL)
    current_row += 1
    if med_picks:
        _write_col_headers(current_row)
        current_row += 1
        med_fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
        for game_label, type_label, rec in med_picks:
            _write_pick_row(current_row, game_label, type_label, rec, med_fill)
            current_row += 1
    else:
        ws.merge_cells(start_row=current_row, start_column=col,
                        end_row=current_row, end_column=col + len(_SUMMARY_HEADERS) - 1)
        cell = ws.cell(row=current_row, column=col,
                        value="No MEDIUM confidence picks today")
        cell.font      = Font(italic=True, color="595959", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")


# ---------------------------------------------------------------------------
# Internal: record sheet writer
# ---------------------------------------------------------------------------

def _write_record_sheet(rs, tally: dict, file_count: int) -> None:
    RECORD_COLS = 6

    def _win_pct(wins: int, losses: int) -> str:
        d = wins + losses
        return f"{wins / d * 100:.1f}%" if d > 0 else "\u2014"

    def _write_section(start_row: int, section_label: str, section_key: str) -> None:
        rs.merge_cells(start_row=start_row, start_column=1,
                        end_row=start_row, end_column=RECORD_COLS)
        sec = rs.cell(row=start_row, column=1, value=section_label)
        sec.fill      = SECTION_FILLS["game"]
        sec.font      = SECTION_FONT
        sec.alignment = Alignment(horizontal="center", vertical="center")
        rs.row_dimensions[start_row].height = 20

        hdr_row = start_row + 1
        for col_i, lbl in enumerate(
            ["BET TYPE", "WINS", "LOSSES", "PUSHES", "TOTAL PICKS", "WIN%"], start=1
        ):
            c = rs.cell(row=hdr_row, column=col_i, value=lbl)
            c.fill      = HEADER_FILL
            c.font      = HEADER_FONT
            c.alignment = Alignment(horizontal="center", vertical="center")
        rs.row_dimensions[hdr_row].height = 16

        total_w = total_l = total_p = 0
        bet_rows = [("ML", "Moneyline"), ("PL", "Puck Line"), ("OU", "Over/Under")]
        for offset, (bet_key, bet_label) in enumerate(bet_rows, start=2):
            w, l, p = tally[section_key][bet_key]
            total_w += w
            total_l += l
            total_p += p
            data = [bet_label, w, l, p, w + l + p, _win_pct(w, l)]
            for col_i, val in enumerate(data, start=1):
                c = rs.cell(row=hdr_row + offset - 1, column=col_i, value=val)
                c.alignment = Alignment(horizontal="center", vertical="center")
            rs.row_dimensions[hdr_row + offset - 1].height = 15

        cmb_row  = hdr_row + len(bet_rows) + 1
        cmb_data = ["COMBINED", total_w, total_l, total_p,
                    total_w + total_l + total_p, _win_pct(total_w, total_l)]
        for col_i, val in enumerate(cmb_data, start=1):
            c = rs.cell(row=cmb_row, column=col_i, value=val)
            c.font      = Font(bold=True)
            c.alignment = Alignment(horizontal="center", vertical="center")
        rs.row_dimensions[cmb_row].height = 15

    rs.merge_cells(start_row=1, start_column=1, end_row=1, end_column=RECORD_COLS)
    title = rs.cell(row=1, column=1, value="MODEL RECORD")
    title.fill      = SECTION_FILLS["game"]
    title.font      = Font(bold=True, color="FFFFFF", size=14)
    title.alignment = Alignment(horizontal="center", vertical="center")
    rs.row_dimensions[1].height = 28

    rs.merge_cells(start_row=2, start_column=1, end_row=2, end_column=RECORD_COLS)
    note = rs.cell(row=2, column=1,
                   value=f"Data from {file_count} prediction file(s)")
    note.font      = Font(italic=True, color="595959", size=9)
    note.alignment = Alignment(horizontal="center", vertical="center")
    rs.row_dimensions[2].height = 14

    _write_section(start_row=4,  section_label="ALL PICKS",            section_key="all")
    _write_section(start_row=11, section_label="HIGH CONFIDENCE ONLY", section_key="high")

    for col_i, w in enumerate([16, 8, 8, 8, 13, 10], start=1):
        rs.column_dimensions[get_column_letter(col_i)].width = w


# ---------------------------------------------------------------------------
# Internal: result evaluation helpers
# ---------------------------------------------------------------------------

def _eval_ml_pick(pick: str, home_abbr: str, away_abbr: str,
                  home_goals: int, away_goals: int):
    """
    Evaluate a moneyline pick. pick format: "CAR ML" or "FLA ML".
    home_abbr / away_abbr: uppercase team abbreviations from the game header.
    """
    if not pick:
        return "", None
    pick_up = pick.upper()
    if home_abbr and pick_up.startswith(home_abbr):
        side = "home"
    elif away_abbr and pick_up.startswith(away_abbr):
        side = "away"
    else:
        return "", None
    result = "WIN" if (side == "home") == (home_goals > away_goals) else "LOSS"
    return result, WIN_FILL if result == "WIN" else LOSS_FILL


def _eval_pl_pick(pick: str, home_abbr: str, away_abbr: str,
                  home_goals: int, away_goals: int):
    """
    Evaluate a puck line pick. pick format: "CAR -1.5" or "FLA +1.5".
    """
    if not pick:
        return "", None
    pick_up = pick.upper()
    margin  = home_goals - away_goals   # positive = home won

    if "-1.5" in pick:
        # Team must win by 2+
        if home_abbr and pick_up.startswith(home_abbr):
            result = "WIN" if margin >= 2 else "LOSS"
        elif away_abbr and pick_up.startswith(away_abbr):
            result = "WIN" if margin <= -2 else "LOSS"
        else:
            return "", None
    elif "+1.5" in pick:
        # Team must not lose by 2+
        if home_abbr and pick_up.startswith(home_abbr):
            result = "WIN" if margin >= -1 else "LOSS"
        elif away_abbr and pick_up.startswith(away_abbr):
            result = "WIN" if margin <= 1 else "LOSS"
        else:
            return "", None
    else:
        return "", None

    return result, WIN_FILL if result == "WIN" else LOSS_FILL


def _eval_ou(pick: str, actual_total: int):
    """Evaluate over/under pick. pick: "OVER 5.5" or "UNDER 5.5"."""
    if not pick:
        return "", None
    m = re.search(r"(\d+\.?\d*)", pick)
    if not m:
        return "", None
    line = float(m.group(1))
    pick_l = pick.upper()
    if abs(actual_total - line) < 0.01:
        return "PUSH", PUSH_FILL
    if "OVER" in pick_l:
        result = "WIN" if actual_total > line else "LOSS"
    elif "UNDER" in pick_l:
        result = "WIN" if actual_total < line else "LOSS"
    else:
        return "", None
    fill = WIN_FILL if result == "WIN" else LOSS_FILL
    return result, fill


# ---------------------------------------------------------------------------
# Internal: layout helpers
# ---------------------------------------------------------------------------

def _set_column_widths(ws) -> None:
    for col_i, width in enumerate(_COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(col_i)].width = width


# ---------------------------------------------------------------------------
# Internal: unit / format helpers
# ---------------------------------------------------------------------------

def _calc_units(result: str, odds_str: str) -> float:
    """Return unit P/L. 1 unit = $10 risked per bet."""
    result = str(result or "").upper()
    if result in ("", "N/A", "PUSH"):
        return 0.0
    if result == "LOSS":
        return -1.0
    if result == "WIN":
        try:
            odds = int(str(odds_str).replace("+", ""))
            return round(odds / 100.0 if odds > 0 else 100.0 / abs(odds), 3)
        except (ValueError, TypeError, ZeroDivisionError):
            return 1.0
    return 0.0


def _fmt_units(units: float) -> str:
    return f"{units:+.2f}u"


def _fmt_score(val) -> str:
    if val is None:
        return ""
    return f"{float(val):.1f}"


def _fmt_pct(val) -> str:
    if val is None:
        return ""
    return f"{float(val) * 100:.1f}%"


def _fmt_odds(val) -> str:
    if val is None:
        return "N/A"
    v = int(val)
    return f"+{v}" if v > 0 else str(v)


def _fmt_edge(val) -> str:
    if val is None:
        return ""
    return f"{float(val):+.1f}%"


def _norm(name: str) -> str:
    return re.sub(r"\s+", "", name.strip().lower())


def _extract_game_label(header_text: str) -> str:
    """Extract 'FLA @ CAR' from the merged header cell (splits on first '|')."""
    return header_text.split("|")[0].strip() if header_text else ""
