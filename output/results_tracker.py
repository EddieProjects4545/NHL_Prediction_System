"""
Results Tracker.

Loads a prior day's recommendations CSV, fetches final NHL scores,
and determines WIN / LOSS / PUSH for each recommendation.

Appends outcomes to results_log.csv and prints a P&L summary.

Usage (via run.py):
    python run.py --track                    # track yesterday
    python run.py --track --date 2026-03-27  # track specific date
"""
import csv
import os
from datetime import date, timedelta
from typing import Dict, List, Optional

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from colorama import Fore, Style

from config import OUTPUT_DIR


# ─── Outcome Logic ────────────────────────────────────────────────────────────

def determine_outcome(rec: Dict, score: Dict) -> str:
    """
    Return 'WIN', 'LOSS', or 'PUSH' for one recommendation given final score.
    rec  : row dict from recommendations CSV
    score: {home_team, away_team, home_goals, away_goals}
    """
    h = int(score.get("home_goals", 0))
    a = int(score.get("away_goals", 0))
    margin = h - a   # positive = home won

    market = rec.get("market", "")
    side   = rec.get("side", "")

    if market == "ML":
        if side == "home":
            return "WIN" if margin > 0 else "LOSS"
        else:
            return "WIN" if margin < 0 else "LOSS"

    elif market == "PL":
        pl_line = rec.get("pl_line", "")
        try:
            line = float(pl_line) if pl_line else None
        except ValueError:
            line = None

        if line is None:
            # Infer from side
            line = -1.5 if side == "home" else 1.5

        if line < 0:          # favourite side (-1.5): must win by 2+
            if side == "home":
                return "WIN" if margin >= 2 else "LOSS"
            else:
                return "WIN" if margin <= -2 else "LOSS"
        else:                 # underdog side (+1.5): must not lose by 2+
            if side == "home":
                return "WIN" if margin >= -1 else "LOSS"
            else:
                return "WIN" if margin <= 1 else "LOSS"

    elif market == "OU":
        total = h + a
        ou_line = rec.get("ou_line", "")
        try:
            line = float(ou_line) if ou_line else 5.5
        except ValueError:
            line = 5.5

        if abs(total - line) < 0.01:
            return "PUSH"
        if side == "over":
            return "WIN" if total > line else "LOSS"
        else:
            return "WIN" if total < line else "LOSS"

    return "UNKNOWN"


def calc_pnl(outcome: str, odds: int, unit: float = 1.0) -> float:
    """Return P&L in units for a 1u bet at American odds."""
    if outcome == "WIN":
        if odds > 0:
            return round(unit * odds / 100, 2)
        else:
            return round(unit * 100 / abs(odds), 2)
    elif outcome == "LOSS":
        return -unit
    return 0.0  # PUSH


# ─── Score Lookup ─────────────────────────────────────────────────────────────

def get_scores_for_date(game_date: str) -> Dict[str, Dict]:
    """
    Return {'{away}@{home}': score_dict} for all completed games on game_date.
    Uses the NHL web API score endpoint (already in nhl_api.py).
    """
    from data.nhl_api import get_game_results_range
    results = get_game_results_range(game_date, game_date)
    index = {}
    for r in results:
        key = f"{r['away_team']}@{r['home_team']}"
        index[key] = r
    return index


# ─── File helpers ─────────────────────────────────────────────────────────────

def load_recommendations(rec_date: str) -> List[Dict]:
    """Load the recommendations CSV for a given date."""
    path = os.path.join(OUTPUT_DIR, f"recommendations_{rec_date}.csv")
    if not os.path.exists(path):
        return []
    with open(path, newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))


# Column layout for results log: (header, width, dict_key)
_LOG_COLS = [
    ("DATE",      12, "date"),
    ("GAME",      22, "game"),
    ("TYPE",       8, "market"),
    ("BET",       30, "bet_label"),
    ("ODDS",       8, "odds"),
    ("EDGE%",      8, "edge_pct"),
    ("CONF",       7, "confidence"),
    ("RESULT",     9, "outcome"),
    ("UNITS",      7, "units"),
    ("P&L",        9, "pnl"),
    ("HOME SCR",   9, "actual_home_goals"),
    ("AWAY SCR",   9, "actual_away_goals"),
]

_TITLE_FILL   = PatternFill(fill_type="solid", fgColor="1F4E79")
_HEADER_FILL  = PatternFill(fill_type="solid", fgColor="2F75B6")
_WIN_FILL     = PatternFill(fill_type="solid", fgColor="C6EFCE")
_LOSS_FILL    = PatternFill(fill_type="solid", fgColor="FFCCCC")
_PUSH_FILL    = PatternFill(fill_type="solid", fgColor="FFEB9C")
_ALT_FILL     = PatternFill(fill_type="solid", fgColor="F2F2F2")
_SUMMARY_FILL = PatternFill(fill_type="solid", fgColor="1F4E79")

_TITLE_FONT   = Font(bold=True, color="FFFFFF", size=13)
_HEADER_FONT  = Font(bold=True, color="FFFFFF", size=10)
_WIN_FONT     = Font(bold=True, color="006100", size=10)
_LOSS_FONT    = Font(bold=True, color="9C0006", size=10)
_PUSH_FONT    = Font(bold=True, color="7D6608", size=10)
_BODY_FONT    = Font(color="000000", size=10)
_SUMMARY_FONT = Font(bold=True, color="FFFFFF", size=10)


def _fmt_log_value(key: str, val) -> str:
    """Format a cell value for the results log."""
    if val is None or val == "":
        return ""
    if key == "edge_pct":
        try: return f"{float(val):+.1f}%"
        except: pass
    elif key == "pnl":
        try: return f"{float(val):+.2f}u"
        except: pass
    elif key == "odds":
        try:
            v = int(val)
            return f"+{v}" if v > 0 else str(v)
        except: pass
    elif key == "units":
        try: return f"{float(val):.1f}u"
        except: pass
    return str(val)


def _migrate_csv_to_excel(ws, next_row: int) -> int:
    """If a legacy results_log.csv exists, import its rows into ws."""
    csv_path = os.path.join(OUTPUT_DIR, "results_log.csv")
    if not os.path.exists(csv_path):
        return next_row
    key_map = {h: k for h, _, k in _LOG_COLS}
    col_keys = [k for _, _, k in _LOG_COLS]
    try:
        with open(csv_path, newline="", encoding="utf-8") as f:
            for row in csv.DictReader(f):
                outcome = str(row.get("outcome", "")).upper()
                fill, font = _row_style(outcome)
                for col_i, key in enumerate(col_keys, start=1):
                    raw = row.get(key, "")
                    cell = ws.cell(row=next_row, column=col_i,
                                   value=_fmt_log_value(key, raw))
                    cell.fill = fill
                    cell.font = font
                    cell.alignment = Alignment(
                        horizontal="left" if col_i == 4 else "center",
                        vertical="center")
                ws.row_dimensions[next_row].height = 16
                next_row += 1
    except Exception:
        pass
    return next_row


def _row_style(outcome: str):
    outcome = outcome.upper()
    if outcome == "WIN":
        return _WIN_FILL, _WIN_FONT
    elif outcome == "LOSS":
        return _LOSS_FILL, _LOSS_FONT
    elif outcome == "PUSH":
        return _PUSH_FILL, _PUSH_FONT
    return _ALT_FILL, _BODY_FONT


_SUMMARY_SENTINEL = "SEASON TOTALS"


def _strip_summary_row(ws) -> None:
    """Remove the summary row if it exists (identified by sentinel in col A)."""
    last = ws.max_row
    if last >= 3 and ws.cell(row=last, column=1).value == _SUMMARY_SENTINEL:
        ws.delete_rows(last)


def _write_summary_row(ws, n_cols: int) -> None:
    """Scan all data rows and write a running-total summary at the bottom."""
    wins = losses = pushes = 0
    total_pnl = 0.0
    total_bets = 0

    # Data starts at row 3 (row 1 = title, row 2 = header)
    result_col = next(i for i, (_, _, k) in enumerate(_LOG_COLS, 1) if k == "outcome")
    pnl_col    = next(i for i, (_, _, k) in enumerate(_LOG_COLS, 1) if k == "pnl")

    for r in range(3, ws.max_row + 1):
        result_val = str(ws.cell(row=r, column=result_col).value or "").upper()
        pnl_raw    = str(ws.cell(row=r, column=pnl_col).value or "").replace("u", "").strip()
        if result_val == "WIN":
            wins += 1
        elif result_val == "LOSS":
            losses += 1
        elif result_val == "PUSH":
            pushes += 1
        else:
            continue
        total_bets += 1
        try:
            total_pnl += float(pnl_raw)
        except ValueError:
            pass

    decided = wins + losses
    win_rate = (wins / decided * 100) if decided > 0 else 0.0
    roi      = (total_pnl / total_bets * 100) if total_bets > 0 else 0.0
    pnl_color = "00B050" if total_pnl >= 0 else "FF0000"

    summary_row = ws.max_row + 1
    ws.row_dimensions[summary_row].height = 18

    # Merge cols 1-4 for label
    ws.merge_cells(start_row=summary_row, start_column=1,
                   end_row=summary_row, end_column=4)
    label_cell = ws.cell(row=summary_row, column=1, value=_SUMMARY_SENTINEL)
    label_cell.fill = _SUMMARY_FILL
    label_cell.font = _SUMMARY_FONT
    label_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Cols 5-7 blank with same fill
    for c in range(5, 8):
        cell = ws.cell(row=summary_row, column=c, value="")
        cell.fill = _SUMMARY_FILL

    # RESULT col: wins / losses / pushes
    r_cell = ws.cell(row=summary_row, column=result_col,
                     value=f"{wins}W / {losses}L / {pushes}P")
    r_cell.fill  = _SUMMARY_FILL
    r_cell.font  = _SUMMARY_FONT
    r_cell.alignment = Alignment(horizontal="center", vertical="center")

    # UNITS col: total bets
    units_col = next(i for i, (_, _, k) in enumerate(_LOG_COLS, 1) if k == "units")
    u_cell = ws.cell(row=summary_row, column=units_col, value=f"{total_bets}u")
    u_cell.fill  = _SUMMARY_FILL
    u_cell.font  = _SUMMARY_FONT
    u_cell.alignment = Alignment(horizontal="center", vertical="center")

    # P&L col
    p_cell = ws.cell(row=summary_row, column=pnl_col,
                     value=f"{total_pnl:+.2f}u")
    p_cell.fill  = _SUMMARY_FILL
    p_cell.font  = Font(bold=True, color=pnl_color, size=10)
    p_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Win rate col (HOME SCR slot)
    wr_col = next(i for i, (_, _, k) in enumerate(_LOG_COLS, 1) if k == "actual_home_goals")
    wr_cell = ws.cell(row=summary_row, column=wr_col,
                      value=f"W%: {win_rate:.1f}%")
    wr_cell.fill  = _SUMMARY_FILL
    wr_cell.font  = _SUMMARY_FONT
    wr_cell.alignment = Alignment(horizontal="center", vertical="center")

    # ROI col (AWAY SCR slot)
    roi_col = next(i for i, (_, _, k) in enumerate(_LOG_COLS, 1) if k == "actual_away_goals")
    roi_cell = ws.cell(row=summary_row, column=roi_col,
                       value=f"ROI: {roi:+.1f}%")
    roi_cell.fill  = _SUMMARY_FILL
    roi_cell.font  = Font(bold=True, color=pnl_color, size=10)
    roi_cell.alignment = Alignment(horizontal="center", vertical="center")


def _row_identity(row_data: Dict) -> tuple:
    """Stable key for one tracked recommendation row."""
    return (
        str(row_data.get("date", "")),
        str(row_data.get("game", "")),
        str(row_data.get("market", "")),
        str(row_data.get("bet_label", "")),
        str(row_data.get("odds", "")),
    )


def _existing_row_keys(ws) -> set[tuple]:
    """Read existing workbook rows into a set of identity tuples."""
    key_cols = {
        "date": next(i for i, (_, _, k) in enumerate(_LOG_COLS, 1) if k == "date"),
        "game": next(i for i, (_, _, k) in enumerate(_LOG_COLS, 1) if k == "game"),
        "market": next(i for i, (_, _, k) in enumerate(_LOG_COLS, 1) if k == "market"),
        "bet_label": next(i for i, (_, _, k) in enumerate(_LOG_COLS, 1) if k == "bet_label"),
        "odds": next(i for i, (_, _, k) in enumerate(_LOG_COLS, 1) if k == "odds"),
    }
    keys = set()
    for r in range(3, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == _SUMMARY_SENTINEL:
            continue
        row_key = (
            str(ws.cell(row=r, column=key_cols["date"]).value or "").strip(),
            str(ws.cell(row=r, column=key_cols["game"]).value or "").strip(),
            str(ws.cell(row=r, column=key_cols["market"]).value or "").strip(),
            str(ws.cell(row=r, column=key_cols["bet_label"]).value or "").strip(),
            str(ws.cell(row=r, column=key_cols["odds"]).value or "").strip(),
        )
        if any(row_key):
            keys.add(row_key)
    return keys


def append_results_log(rows: List[Dict]) -> str:
    """Append tracked outcomes to results_log.xlsx. Returns path."""
    path = os.path.join(OUTPUT_DIR, "results_log.xlsx")
    n_cols = len(_LOG_COLS)

    if os.path.exists(path):
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        _strip_summary_row(ws)
        next_row = ws.max_row + 1
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Results Log"

        # Title row
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
        tc = ws.cell(row=1, column=1, value="NHL Betting — Results Log")
        tc.fill = _TITLE_FILL
        tc.font = _TITLE_FONT
        tc.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 26

        # Header row
        for col_i, (header, width, _) in enumerate(_LOG_COLS, start=1):
            hc = ws.cell(row=2, column=col_i, value=header)
            hc.fill = _HEADER_FILL
            hc.font = _HEADER_FONT
            hc.alignment = Alignment(horizontal="center", vertical="center")
            ws.column_dimensions[get_column_letter(col_i)].width = width
        ws.row_dimensions[2].height = 18
        ws.freeze_panes = "A3"

        next_row = _migrate_csv_to_excel(ws, 3)

    existing_keys = _existing_row_keys(ws)

    col_keys = [k for _, _, k in _LOG_COLS]
    for row_data in rows:
        row_key = _row_identity(row_data)
        if row_key in existing_keys:
            continue
        outcome = str(row_data.get("outcome", "")).upper()
        fill, font = _row_style(outcome)
        for col_i, key in enumerate(col_keys, start=1):
            raw = row_data.get(key, "")
            cell = ws.cell(row=next_row, column=col_i,
                           value=_fmt_log_value(key, raw))
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(
                horizontal="left" if col_i == 4 else "center",
                vertical="center")
        ws.row_dimensions[next_row].height = 16
        next_row += 1
        existing_keys.add(row_key)

    _write_summary_row(ws, n_cols)
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    wb.save(path)
    return path


# ─── Main entry point ─────────────────────────────────────────────────────────

def track_results(rec_date: Optional[str] = None) -> None:
    """
    Load recommendations for rec_date, fetch scores, compute outcomes,
    print P&L summary, and append to results_log.csv.
    """
    if rec_date is None:
        rec_date = (date.today() - timedelta(days=1)).isoformat()

    print(Fore.CYAN + Style.BRIGHT + f"\n  RESULTS TRACKING — {rec_date}")
    print(Fore.WHITE + "  " + "─" * 61)

    recs = load_recommendations(rec_date)
    if not recs:
        print(Fore.YELLOW +
              f"  No recommendations file found for {rec_date}.")
        print(Fore.WHITE +
              f"  Expected: {os.path.join(OUTPUT_DIR, f'recommendations_{rec_date}.csv')}\n")
        return

    scores = get_scores_for_date(rec_date)
    if not scores:
        print(Fore.YELLOW +
              f"  No completed game scores found for {rec_date}. "
              "Games may not have finished yet.\n")
        return

    tracked_rows = []
    wins = losses = pushes = skipped = 0
    total_pnl = 0.0
    total_bet = 0.0

    # Header
    print(Fore.WHITE + Style.BRIGHT +
          f"  {'#':<4} {'Bet':<35} {'Odds':>6} {'Result':>6} {'P&L':>9}")
    print(Fore.WHITE + "  " + "─" * 61)

    for rec in recs:
        game_str = rec.get("game", "")             # "ANA @ EDM"
        parts = game_str.split(" @ ")
        if len(parts) != 2:
            skipped += 1
            continue
        away_abbr, home_abbr = parts[0].strip(), parts[1].strip()
        score_key = f"{away_abbr}@{home_abbr}"

        if score_key not in scores:
            skipped += 1
            continue

        score   = scores[score_key]
        outcome = determine_outcome(rec, score)
        try:
            odds = int(rec.get("odds", 0))
        except (ValueError, TypeError):
            skipped += 1
            continue

        pnl = calc_pnl(outcome, odds)
        total_pnl += pnl
        total_bet += 1.0

        if outcome == "WIN":
            wins += 1
            result_color = Fore.GREEN
        elif outcome == "LOSS":
            losses += 1
            result_color = Fore.RED
        elif outcome == "PUSH":
            pushes += 1
            result_color = Fore.YELLOW
        else:
            skipped += 1
            continue

        rank     = rec.get("rank", "?")
        bet_lbl  = rec.get("bet_label", rec.get("market", ""))[:34]
        odds_str = f"+{odds}" if odds > 0 else str(odds)
        pnl_str  = f"{pnl:+.2f}"

        print(
            Fore.WHITE + f"  #{rank:<3} {bet_lbl:<35} {odds_str:>6} " +
            result_color + f"{outcome:>6}" +
            (Fore.GREEN if pnl >= 0 else Fore.RED) + f" {pnl_str:>8}u"
        )

        tracked_rows.append({
            "date"              : rec_date,
            "game"              : game_str,
            "market"            : rec.get("market", ""),
            "side"              : rec.get("side", ""),
            "bet_label"         : rec.get("bet_label", ""),
            "odds"              : odds,
            "model_prob"        : rec.get("model_prob", ""),
            "edge_pct"          : rec.get("edge_pct", ""),
            "confidence"        : rec.get("confidence", ""),
            "outcome"           : outcome,
            "units"             : 1.0,
            "pnl"               : pnl,
            "actual_home_goals" : score["home_goals"],
            "actual_away_goals" : score["away_goals"],
        })

    # ── Summary ───────────────────────────────────────────────────────────────
    total_decided = wins + losses + pushes
    print(Fore.WHITE + "  " + "─" * 61)

    if total_decided == 0:
        print(Fore.YELLOW + "  No completed games matched recommendations.\n")
        return

    roi = (total_pnl / total_bet * 100) if total_bet > 0 else 0.0
    acc = (wins / (wins + losses) * 100) if (wins + losses) > 0 else 0.0

    pnl_color = Fore.GREEN if total_pnl >= 0 else Fore.RED
    print(
        Fore.WHITE + Style.BRIGHT +
        f"  Bets: {total_decided}  |  "
        f"Won: {wins}  Lost: {losses}  Push: {pushes}  Skipped: {skipped}"
    )
    print(
        pnl_color + Style.BRIGHT +
        f"  P&L (@ 1u/bet): {total_pnl:+.2f}u  |  "
        f"ROI: {roi:+.1f}%  |  Win rate: {acc:.1f}%"
    )

    if tracked_rows:
        log_path = append_results_log(tracked_rows)
        print(Fore.WHITE + f"  Appended {len(tracked_rows)} records → {log_path}")

    print()


def track_and_update_excel(rec_date: Optional[str] = None,
                            output_dir: Optional[str] = None) -> None:
    """
    Run the normal CSV-based results tracking, then write results back into
    the Excel predictions workbook for rec_date (if it exists).
    """
    if rec_date is None:
        rec_date = (date.today() - timedelta(days=1)).isoformat()

    # Run existing terminal / CSV tracking
    track_results(rec_date)

    # Write back into Excel workbook
    scores = get_scores_for_date(rec_date)
    if not scores:
        return

    completed = []
    for score_key, score in scores.items():
        parts = score_key.split("@")
        if len(parts) != 2:
            continue
        away_abbr, home_abbr = parts[0], parts[1]
        completed.append({
            "home_name":  home_abbr,
            "away_name":  away_abbr,
            "home_score": score.get("home_goals", 0),
            "away_score": score.get("away_goals", 0),
        })

    if completed:
        from output.excel_writer import write_results_to_workbook
        write_results_to_workbook(rec_date, completed, output_dir)
        print(Fore.GREEN + f"  Excel write-back complete: nhl_predictions_{rec_date}.xlsx\n")
